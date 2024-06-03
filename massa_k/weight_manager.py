import os
import time
from datetime import datetime
from threading import Thread

from openpyxl.reader.excel import load_workbook

from massa_k.massa_k_scales import MassaKScales
from openpyxl import Workbook


class WeightManager(Thread):

    SEND_INTERVAL = 0.1  # seconds

    def __init__(self, host, port, data_folder_path):
        super().__init__()
        self.__host = host
        self.__port = port
        self.__data_folder_path = data_folder_path
        self.__is_started = False
        self.__cur_file_path = ''
        self.__workbook = None
        self.__sheet = None
        self.__next_id = 0

        self._scales = MassaKScales()
        self._scales.start()

        self.check_change_day()

    def check_change_day(self):
        # Create data folder if not exist
        if not os.path.exists(self.__data_folder_path):
            os.makedirs(self.__data_folder_path)

        cur_date_time_str = datetime.now().strftime('%d-%m-%Y')
        new_file_path = os.path.join(self.__data_folder_path, cur_date_time_str + '.xlsx')

        if self.__cur_file_path != new_file_path:
            if self.__workbook is not None:
                self.__workbook.save(self.__cur_file_path)

            self.__cur_file_path = new_file_path

            if os.path.exists(self.__cur_file_path):
                self.__workbook = load_workbook(self.__cur_file_path)
                self.__sheet = self.__workbook.active
                self.__next_id = self.__sheet.max_row + 1
            else:
                self.__workbook = Workbook()
                self.__sheet = self.__workbook.active
                self.__next_id = 1

    def weight_event_handler(self, mass):
        self.check_change_day()
        current_time_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        self.__sheet.append([self.__next_id, current_time_str, float(mass)])
        self.__next_id += 1
        self.__workbook.save(self.__cur_file_path)

        print(current_time_str, " Масса:", mass, "г.")

    def run(self) -> None:

        self.__is_started = True

        while self.__is_started:

            if not self._scales.is_connected:

                print("Подключение к весам ", self.__host, "(port", self.__port, ")...")
                self._scales.connect(self.__host, self.__port)

                if self._scales.is_connected:
                    print("Подключено!")
                    self._scales.add_weight_event_handler(self.weight_event_handler)
                    # self._scales.start()
                else:
                    print("Ошибка соединения с весами! IP: ", self.__host, " , port: ", self.__port)
            else:
                self._scales.send_command()
                time.sleep(self.SEND_INTERVAL)

    def stop(self):
        self.__is_started = False
